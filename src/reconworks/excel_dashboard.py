from __future__ import annotations

from pathlib import Path
from typing import Dict, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from .config import ProjectConfig
from .db import (
    connect,
    latest_batch_id,
    create_excel_runs_table,
    insert_excel_run,
    delete_where_batch,
)
from .util import utc_now_iso, ensure_dir


def _autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row, 200) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 50)


def build_excel(repo_root: Path, cfg: ProjectConfig, batch_id: Optional[str] = None) -> Dict[str, str]:
    out_dir = repo_root / cfg.output_dir
    ensure_dir(out_dir / "excel")

    conn = connect(repo_root / cfg.database_path)
    create_excel_runs_table(conn)

    b = batch_id or latest_batch_id(conn)
    if not b:
        conn.close()
        return {"output_path": ""}

    # Pull data
    exceptions = pd.read_sql_query(
        "SELECT * FROM exceptions WHERE batch_id=? ORDER BY severity, exception_code",
        conn, params=(b,)
    )
    matches = pd.read_sql_query(
        "SELECT * FROM matches WHERE batch_id=? ORDER BY match_score DESC",
        conn, params=(b,)
    )
    qa = pd.read_sql_query(
        "SELECT * FROM qa_flags WHERE batch_id=? ORDER BY severity, flag_code",
        conn, params=(b,)
    )
    spend = pd.read_sql_query(
        "SELECT * FROM rpt_spend_by_month_vendor WHERE batch_id=?",
        conn, params=(b,)
    )
    match_rate = pd.read_sql_query(
        "SELECT * FROM rpt_match_rate_by_month WHERE batch_id=? ORDER BY month",
        conn, params=(b,)
    )
    top_vendors = pd.read_sql_query(
        "SELECT * FROM rpt_top_vendors WHERE batch_id=? ORDER BY spend_usd DESC",
        conn, params=(b,)
    )

    output_path = repo_root / cfg.excel.output_path

    # Write base sheets via pandas
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame([{"batch_id": b, "generated_at_utc": utc_now_iso()}]).to_excel(writer, sheet_name="Summary", index=False)
        exceptions.to_excel(writer, sheet_name="Exceptions", index=False)
        matches.to_excel(writer, sheet_name="Matches", index=False)
        qa.to_excel(writer, sheet_name="QAFlags", index=False)
        spend.to_excel(writer, sheet_name="SpendByVendorMonth", index=False)
        match_rate.to_excel(writer, sheet_name="MatchRateByMonth", index=False)
        top_vendors.to_excel(writer, sheet_name="TopVendors", index=False)

    wb = load_workbook(output_path)
    ws = wb["Summary"]

    # Dashboard-ish formatting
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws["A1"] = "ReconWorks Dashboard"
    ws["A1"].font = Font(size=18, bold=True)

    # Header block
    ws["A3"] = "Batch ID"
    ws["B3"] = b
    ws["A4"] = "Generated (UTC)"
    ws["B4"] = utc_now_iso()

    # KPI block (latest month)
    latest_month = None
    if not match_rate.empty and "month" in match_rate.columns:
        latest_month = str(match_rate["month"].iloc[-1])

    total_tx = int(match_rate["total_transactions"].iloc[-1]) if (not match_rate.empty and "total_transactions" in match_rate.columns) else 0
    matched_tx = int(match_rate["matched_transactions"].iloc[-1]) if (not match_rate.empty and "matched_transactions" in match_rate.columns) else len(matches)
    mr = float(match_rate["match_rate"].iloc[-1]) if (not match_rate.empty and "match_rate" in match_rate.columns) else (matched_tx / total_tx if total_tx else 0.0)

    ws["A6"] = "Latest Month"
    ws["B6"] = latest_month or ""
    ws["A7"] = "Transactions"
    ws["B7"] = total_tx
    ws["A8"] = "Matched"
    ws["B8"] = matched_tx
    ws["A9"] = "Match Rate"
    ws["B9"] = mr
    ws["B9"].number_format = "0.00%"

    ws["A11"] = "Exceptions"
    ws["B11"] = int(len(exceptions)) if exceptions is not None else 0

    for r in range(3, 12):
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"A{r}"].alignment = Alignment(horizontal="left")
        ws[f"B{r}"].alignment = Alignment(horizontal="left")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 50

    # Small Exceptions preview (top 10)
    ws["A13"] = "Exceptions (Top 10)"
    ws["A13"].font = Font(bold=True)
    preview_cols = ["severity", "exception_code", "record_type", "message"]
    preview = exceptions[preview_cols].head(10) if (exceptions is not None and not exceptions.empty) else pd.DataFrame(columns=preview_cols)
    start_row = 14
    for j, col in enumerate(preview_cols, start=1):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.font = Font(bold=True)
    for i, row in enumerate(preview.itertuples(index=False), start=1):
        for j, val in enumerate(row, start=1):
            ws.cell(row=start_row + i, column=j, value=val)
    # Set widths for preview table
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 60

    # Chart 1: Match rate by month (use column chart so a single month still looks good)
    if "MatchRateByMonth" in wb.sheetnames:
        mws = wb["MatchRateByMonth"]
        headers = [c.value for c in mws[1]]
        if "month" in headers and "match_rate" in headers and mws.max_row >= 2:
            month_col = headers.index("month") + 1
            rate_col = headers.index("match_rate") + 1

            chart = BarChart()
            chart.type = "col"
            chart.title = "Match Rate by Month"
            chart.y_axis.number_format = "0%"
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 1
            chart.y_axis.majorGridlines = None  # reduce visual noise (Excel will render without heavy lines)
            chart.legend = None

            data = Reference(mws, min_col=rate_col, min_row=1, max_row=mws.max_row)  # includes header
            cats = Reference(mws, min_col=month_col, min_row=2, max_row=mws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 7
            chart.width = 18
            ws.add_chart(chart, "D3")

    # Chart 2: Top vendors by spend
    if "TopVendors" in wb.sheetnames:
        tws = wb["TopVendors"]
        headers = [c.value for c in tws[1]]
        if "vendor_canonical" in headers and "spend_usd" in headers and tws.max_row >= 2:
            vcol = headers.index("vendor_canonical") + 1
            scol = headers.index("spend_usd") + 1

            chart = BarChart()
            chart.type = "col"
            chart.title = "Top Vendors by Spend (USD)"
            chart.y_axis.number_format = "$#,##0"
            chart.y_axis.majorGridlines = None
            chart.legend = None

            data = Reference(tws, min_col=scol, min_row=1, max_row=min(tws.max_row, 21))
            cats = Reference(tws, min_col=vcol, min_row=2, max_row=min(tws.max_row, 21))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 7
            chart.width = 18
            ws.add_chart(chart, "D18")

    # Autosize key sheets
    for name in ["Summary","Exceptions","Matches","QAFlags","SpendByVendorMonth","MatchRateByMonth","TopVendors"]:
        if name in wb.sheetnames:
            _autosize(wb[name])

    wb.save(output_path)

    delete_where_batch(conn, "excel_runs", b)
    insert_excel_run(conn, {"created_at_utc": utc_now_iso(), "batch_id": b, "output_path": str(cfg.excel.output_path)})
    conn.close()
    return {"output_path": str(output_path)}
